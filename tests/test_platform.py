"""Phase 12: notifications, Excel import/export.

The definition of done, restated:

* imports never write invalid rows,
* audit covers every listed action,
* notifications dedupe so a fact alerts once.

The import tests carry the most weight. A spreadsheet is the least trustworthy
input the platform accepts, and "we imported 400 rows and 30 were wrong" is how
a clean database stops being clean.
"""

from __future__ import annotations

import csv
import io
import uuid
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal

import pytest
from openpyxl import Workbook, load_workbook

from app.core.permissions import Permission, Role, permissions_for
from app.engines.importing.schema import (
    SCHEMAS,
    CoercionError,
    Column,
    coerce,
    normalise_header,
    schema_for,
)
from app.engines.importing.workbook import WorkbookError, parse, write_template
from app.models.platform import ImportEntity
from app.services.notifications import SLA_MILESTONES, sla_milestone

API = "/api/v1"


def csv_bytes(headers: list[str], rows: list[list[object]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return buffer.getvalue().encode()


def xlsx_bytes(headers: list[str], rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# ----------------------------------------------------------------- coercion


class TestHeaderMatching:
    @pytest.mark.parametrize(
        "raw", ["Full Name", "full_name", "FULL NAME", " Full-Name ", "full  name"]
    )
    def test_headers_match_however_a_user_typed_them(self, raw):
        # Users should not have to match our spelling to import their own data.
        assert normalise_header(raw) == "full_name"

    def test_an_empty_header_normalises_to_nothing(self):
        assert normalise_header(None) == ""
        assert normalise_header("   ") == ""


class TestCoercion:
    def test_a_required_empty_cell_is_an_error_naming_the_column(self):
        with pytest.raises(CoercionError, match="Full name is required"):
            coerce(None, Column("full_name", "Full name", required=True))

    def test_an_optional_empty_cell_is_none_not_an_error(self):
        assert coerce("   ", Column("notes", "Notes")) is None

    @pytest.mark.parametrize("raw", ["yes", "Y", "TRUE", "1", "t"])
    def test_every_way_a_human_writes_yes(self, raw):
        assert coerce(raw, Column("flag", "Flag", "bool")) is True

    @pytest.mark.parametrize("raw", ["no", "N", "FALSE", "0"])
    def test_every_way_a_human_writes_no(self, raw):
        assert coerce(raw, Column("flag", "Flag", "bool")) is False

    def test_an_unreadable_boolean_is_reported_not_guessed(self):
        with pytest.raises(CoercionError, match="must be yes or no"):
            coerce("maybe", Column("flag", "Flag", "bool"))

    def test_money_survives_currency_symbols_and_separators(self):
        column = Column("rate", "Rate", "decimal")
        assert coerce("QAR 22,000.50", column) == Decimal("22000.50")
        assert coerce(" 18000 ", column) == Decimal("18000")

    def test_a_negative_rate_is_rejected(self):
        with pytest.raises(CoercionError, match="cannot be negative"):
            coerce("-500", Column("rate", "Rate", "decimal"))

    @pytest.mark.parametrize(
        "raw", ["2026-03-15", "15/03/2026", "15-03-2026", "15 Mar 2026", "2026/03/15"]
    )
    def test_dates_are_read_in_the_formats_people_actually_use(self, raw):
        assert coerce(raw, Column("d", "Date", "date")) == date(2026, 3, 15)

    def test_a_real_datetime_passes_straight_through(self):
        assert coerce(datetime(2026, 3, 15, 9, 0), Column("d", "Date", "date")) == date(2026, 3, 15)

    def test_an_unreadable_date_is_reported_rather_than_guessed(self):
        with pytest.raises(CoercionError, match="not a date we can read"):
            coerce("next Tuesday", Column("d", "Date", "date"))

    def test_a_choice_outside_the_list_names_the_options(self):
        column = Column("t", "Type", "choice", choices=("CUSTOMER", "PARTNER"))
        assert coerce("customer", column) == "CUSTOMER"
        with pytest.raises(CoercionError, match="CUSTOMER, PARTNER"):
            coerce("supplier", column)

    def test_a_country_must_be_two_letters(self):
        column = Column("c", "Country", "country")
        assert coerce("qa", column) == "QA"
        with pytest.raises(CoercionError, match="two-letter"):
            coerce("Qatar", column)

    def test_an_invalid_email_is_rejected(self):
        column = Column("e", "Email", "email")
        assert coerce(" Person@Example.COM ", column) == "person@example.com"
        with pytest.raises(CoercionError, match="not a valid email"):
            coerce("not-an-email", column)

    def test_a_list_splits_on_the_usual_separators(self):
        column = Column("s", "Skills", "list")
        assert coerce("SAP FICO; Power BI, Java", column) == ["SAP FICO", "Power BI", "Java"]

    def test_text_length_is_enforced(self):
        with pytest.raises(CoercionError, match="longer than 5"):
            coerce("far too long", Column("t", "Title", max_length=5))


class TestWorkbookParsing:
    def test_a_csv_and_an_xlsx_parse_identically(self):
        schema = schema_for(ImportEntity.CUSTOMERS)
        headers = ["Name", "Country"]
        rows = [["Milaha", "QA"]]

        from_csv = parse(csv_bytes(headers, rows), filename="a.csv", schema=schema)
        from_xlsx = parse(xlsx_bytes(headers, rows), filename="a.xlsx", schema=schema)

        assert from_csv.rows == from_xlsx.rows

    def test_the_first_data_row_is_numbered_two_as_excel_shows_it(self):
        schema = schema_for(ImportEntity.CUSTOMERS)
        sheet = parse(csv_bytes(["Name"], [["A"], ["B"]]), filename="a.csv", schema=schema)
        assert [number for number, _ in sheet.rows] == [2, 3]

    def test_a_missing_required_column_is_reported_at_file_level(self):
        schema = schema_for(ImportEntity.CUSTOMERS)
        sheet = parse(csv_bytes(["Country"], [["QA"]]), filename="a.csv", schema=schema)
        assert "Name" in sheet.missing_required

    def test_unknown_columns_are_reported_not_silently_dropped(self):
        schema = schema_for(ImportEntity.CUSTOMERS)
        sheet = parse(
            csv_bytes(["Name", "Favourite Colour"], [["Milaha", "blue"]]),
            filename="a.csv",
            schema=schema,
        )
        assert sheet.unknown_headers == ["Favourite Colour"]

    def test_an_empty_file_is_rejected(self):
        with pytest.raises(WorkbookError, match="empty"):
            parse(b"", filename="a.csv", schema=schema_for(ImportEntity.CUSTOMERS))

    def test_an_unsupported_extension_is_rejected(self):
        with pytest.raises(WorkbookError, match=r"\.xlsx or \.csv"):
            parse(b"data", filename="a.pdf", schema=schema_for(ImportEntity.CUSTOMERS))

    def test_a_template_round_trips_through_the_parser(self):
        """The template we hand out must be a file we can read back."""
        schema = schema_for(ImportEntity.RESOURCES)
        sheet = parse(write_template(schema), filename="t.xlsx", schema=schema)
        assert sheet.missing_required == []

    def test_every_schema_has_an_identity_rule(self):
        # Without one, duplicate detection silently does nothing.
        for entity, schema in SCHEMAS.items():
            assert schema.identity_fields, f"{entity.value} has no identity rule"


# ------------------------------------------------------------------ imports


@pytest.fixture
async def importer(as_role):
    client, user = await as_role(Role.SALES)
    return client, user


async def upload(client, entity: str, payload: bytes, filename: str = "data.csv"):
    return await client.post(
        f"{API}/imports/{entity}/upload",
        files={"file": (filename, payload, "text/csv")},
    )


class TestImportStaging:
    async def test_upload_stages_without_writing_anything(self, importer):
        client, _ = importer
        name = f"Import Co {uuid.uuid4().hex[:8]}"

        response = await upload(client, "customers", csv_bytes(["Name", "Country"], [[name, "QA"]]))
        assert response.status_code == 201, response.text
        body = response.json()

        assert body["batch"]["status"] == "STAGED"
        assert body["batch"]["valid_rows"] + body["batch"]["warning_rows"] == 1

        # Nothing in the business table yet.
        accounts = (await client.get(f"{API}/accounts?q={name}")).json()
        assert accounts["total"] == 0

    async def test_an_invalid_row_is_classified_not_rejected_wholesale(self, importer):
        client, _ = importer
        payload = csv_bytes(
            ["Name", "Country", "Account type"],
            [
                [f"Good {uuid.uuid4().hex[:6]}", "QA", "CUSTOMER"],
                ["", "QA", "CUSTOMER"],
                [f"Bad Type {uuid.uuid4().hex[:6]}", "QA", "SUPPLIER"],
            ],
        )

        body = (await upload(client, "customers", payload)).json()
        assert body["batch"]["total_rows"] == 3
        assert body["batch"]["invalid_rows"] == 2

        invalid = [row for row in body["rows"] if row["validation_state"] == "INVALID"]
        assert len(invalid) == 2
        # Every error names the column so the user can fix it.
        assert any("Name is required" in error for row in invalid for error in row["errors"])
        assert any("Account type" in error for row in invalid for error in row["errors"])

    async def test_the_error_names_the_row_number_the_user_can_see(self, importer):
        client, _ = importer
        payload = csv_bytes(
            ["Name", "Country"], [[f"Fine {uuid.uuid4().hex[:6]}", "QA"], ["", "AE"]]
        )

        body = (await upload(client, "customers", payload)).json()
        bad = next(row for row in body["rows"] if row["validation_state"] == "INVALID")
        # Row 1 is the header, so the second data row is row 3 in Excel.
        assert bad["row_number"] == 3

    async def test_a_missing_required_column_fails_the_file_once(self, importer):
        client, _ = importer
        body = (await upload(client, "customers", csv_bytes(["Country"], [["QA"], ["AE"]]))).json()

        assert body["batch"]["invalid_rows"] == 2
        assert any("Name" in error for error in body["batch"]["file_errors"])

    async def test_a_within_file_duplicate_is_flagged(self, importer):
        client, _ = importer
        name = f"Twin {uuid.uuid4().hex[:8]}"
        body = (
            await upload(
                client, "customers", csv_bytes(["Name", "Country"], [[name, "QA"], [name, "QA"]])
            )
        ).json()

        assert body["batch"]["duplicate_rows"] == 1

    async def test_an_advisory_does_not_block_a_row(self, importer):
        client, _ = importer
        # No country: importable, but duplicate detection is weaker.
        body = (
            await upload(
                client, "customers", csv_bytes(["Name"], [[f"Advisory {uuid.uuid4().hex[:6]}"]])
            )
        ).json()

        row = body["rows"][0]
        assert row["validation_state"] == "WARNING"
        assert row["warnings"]
        assert body["batch"]["is_committable"] is True

    async def test_a_wholly_blank_row_is_skipped_not_reported_as_invalid(self, importer):
        """Trailing blank rows are ubiquitous in Excel and mean nothing."""
        client, _ = importer
        payload = csv_bytes(
            ["Name", "Country"],
            [[f"Real {uuid.uuid4().hex[:6]}", "QA"], ["", ""], ["", ""]],
        )

        body = (await upload(client, "customers", payload)).json()
        assert body["batch"]["total_rows"] == 1
        assert body["batch"]["invalid_rows"] == 0

    async def test_an_unreadable_file_is_a_422(self, importer):
        client, _ = importer
        response = await client.post(
            f"{API}/imports/customers/upload",
            files={"file": ("data.pdf", b"%PDF-1.4", "application/pdf")},
        )
        assert response.status_code == 422

    async def test_staging_is_audited(self, importer, as_role):
        client, _ = importer
        await upload(
            client, "customers", csv_bytes(["Name"], [[f"Audited {uuid.uuid4().hex[:6]}"]])
        )

        admin, _ = await as_role(Role.ADMIN)
        entries = (await admin.get(f"{API}/audit?action=IMPORT_STAGED")).json()
        assert entries["items"]


class TestImportCommit:
    async def test_committing_writes_only_the_valid_rows(self, importer):
        """The Phase 12 definition of done."""
        client, _ = importer
        good = f"Committed {uuid.uuid4().hex[:8]}"
        payload = csv_bytes(
            ["Name", "Country", "Account type"],
            [[good, "QA", "CUSTOMER"], ["", "QA", "CUSTOMER"]],
        )

        batch = (await upload(client, "customers", payload)).json()["batch"]
        result = (await client.post(f"{API}/imports/{batch['id']}/commit")).json()

        assert result["created"] == 1
        assert result["never_written"] == 1

        accounts = (await client.get(f"{API}/accounts?q={good}")).json()
        assert accounts["total"] == 1

    async def test_an_invalid_row_never_reaches_the_database(self, importer):
        client, _ = importer
        marker = uuid.uuid4().hex[:8]
        payload = csv_bytes(
            ["Name", "Country"],
            [[f"Valid {marker}", "QA"], [f"Bad {marker}", "Qatar"]],
        )

        batch = (await upload(client, "customers", payload)).json()["batch"]
        await client.post(f"{API}/imports/{batch['id']}/commit")

        found = (await client.get(f"{API}/accounts?q={marker}")).json()
        assert found["total"] == 1
        assert found["items"][0]["name"] == f"Valid {marker}"

    async def test_an_existing_record_is_skipped_not_duplicated(self, importer):
        client, _ = importer
        name = f"Existing {uuid.uuid4().hex[:8]}"

        first = (
            await upload(client, "customers", csv_bytes(["Name", "Country"], [[name, "QA"]]))
        ).json()["batch"]
        await client.post(f"{API}/imports/{first['id']}/commit")

        # Re-importing last month's spreadsheet is a common accident.
        second = (
            await upload(client, "customers", csv_bytes(["Name", "Country"], [[name, "QA"]]))
        ).json()["batch"]
        assert second["duplicate_rows"] == 1

        response = await client.post(f"{API}/imports/{second['id']}/commit")
        assert response.status_code == 422

        found = (await client.get(f"{API}/accounts?q={name}")).json()
        assert found["total"] == 1

    async def test_committing_twice_is_refused(self, importer):
        client, _ = importer
        batch = (
            await upload(
                client, "customers", csv_bytes(["Name"], [[f"Once {uuid.uuid4().hex[:6]}"]])
            )
        ).json()["batch"]

        assert (await client.post(f"{API}/imports/{batch['id']}/commit")).status_code == 200
        assert (await client.post(f"{API}/imports/{batch['id']}/commit")).status_code == 409

    async def test_a_batch_of_only_invalid_rows_cannot_be_committed(self, importer):
        client, _ = importer
        batch = (
            await upload(client, "customers", csv_bytes(["Name", "Country"], [["", "QA"]]))
        ).json()["batch"]
        assert batch["invalid_rows"] == 1

        response = await client.post(f"{API}/imports/{batch['id']}/commit")
        assert response.status_code == 422
        assert "nothing to import" in response.text.lower()

    async def test_a_discarded_batch_cannot_be_committed(self, importer):
        client, _ = importer
        batch = (
            await upload(
                client, "customers", csv_bytes(["Name"], [[f"Gone {uuid.uuid4().hex[:6]}"]])
            )
        ).json()["batch"]

        await client.post(f"{API}/imports/{batch['id']}/discard")
        assert (await client.post(f"{API}/imports/{batch['id']}/commit")).status_code == 409

    async def test_resources_import_with_their_skills(self, importer, as_role):
        client, _ = importer
        name = f"Imported Consultant {uuid.uuid4().hex[:8]}"
        payload = csv_bytes(
            ["Full name", "Resource type", "Skills", "Country"],
            [[name, "CONSULTANT", "SAP FICO; Power BI", "QA"]],
        )

        batch = (await upload(client, "resources", payload)).json()["batch"]
        result = (await client.post(f"{API}/imports/{batch['id']}/commit")).json()
        assert result["created"] == 1

        hr, _ = await as_role(Role.HR_RESOURCING)
        found = (await hr.get(f"{API}/resources?q={name}")).json()
        assert found["total"] == 1
        assert len(found["items"][0]["skills"]) == 2

    async def test_committing_is_audited(self, importer, as_role):
        client, _ = importer
        batch = (
            await upload(
                client, "customers", csv_bytes(["Name"], [[f"Logged {uuid.uuid4().hex[:6]}"]])
            )
        ).json()["batch"]
        await client.post(f"{API}/imports/{batch['id']}/commit")

        admin, _ = await as_role(Role.ADMIN)
        entries = (await admin.get(f"{API}/audit?action=IMPORT_COMMITTED")).json()
        assert entries["items"]


class TestImportArtifacts:
    async def test_the_template_is_a_readable_workbook(self, importer):
        client, _ = importer
        response = await client.get(f"{API}/imports/customers/template.xlsx")

        assert response.status_code == 200
        workbook = load_workbook(io.BytesIO(response.content))
        assert workbook.active["A1"].value == "Name*"

    async def test_the_error_report_carries_the_original_columns(self, importer):
        client, _ = importer
        batch = (
            await upload(
                client,
                "customers",
                csv_bytes(["Name", "Country"], [["", "QA"]]),
            )
        ).json()["batch"]

        response = await client.get(f"{API}/imports/{batch['id']}/errors.xlsx")
        assert response.status_code == 200

        sheet = load_workbook(io.BytesIO(response.content)).active
        headers = [cell.value for cell in sheet[1]]
        assert headers[:2] == ["Row", "Problem"]
        # The original columns are kept so the fixed file can be re-uploaded.
        assert "Name" in headers
        assert sheet[2][1].value and "required" in sheet[2][1].value.lower()

    async def test_the_entity_catalogue_describes_every_column(self, importer):
        client, _ = importer
        entities = (await client.get(f"{API}/imports/entities")).json()

        assert {item["entity"] for item in entities} >= {"customers", "resources"}
        for item in entities:
            assert item["columns"]
            assert all("label" in column for column in item["columns"])


class TestImportAuthorization:
    async def test_management_cannot_import(self, as_role):
        client, _ = await as_role(Role.MANAGEMENT)
        response = await upload(client, "customers", csv_bytes(["Name"], [["Nope"]]))
        assert response.status_code == 403


# ------------------------------------------------------------------ exports


class TestExports:
    async def test_an_export_is_a_readable_workbook(self, as_role):
        client, _ = await as_role(Role.SALES)
        response = await client.get(f"{API}/exports/customers.xlsx")

        assert response.status_code == 200
        assert "attachment" in response.headers["content-disposition"]
        sheet = load_workbook(io.BytesIO(response.content)).active
        assert [cell.value for cell in sheet[1]][:2] == ["Name", "Account type"]

    async def test_sales_export_omits_the_consultant_cost_column(self, as_role):
        """An export that ignored RBAC would be the easiest bypass in the system."""
        client, _ = await as_role(Role.SALES)
        assert Permission.FIELD_RESOURCE_COST not in permissions_for(Role.SALES)

        sheet = load_workbook(
            io.BytesIO((await client.get(f"{API}/exports/resources.xlsx")).content)
        ).active
        headers = [cell.value for cell in sheet[1]]

        # Omitted entirely, not blanked — an empty cell would be ambiguous.
        assert "Expected cost" not in headers

    async def test_resourcing_export_includes_cost_but_not_the_bill_rate(self, as_role):
        client, _ = await as_role(Role.HR_RESOURCING)

        resources = load_workbook(
            io.BytesIO((await client.get(f"{API}/exports/resources.xlsx")).content)
        ).active
        assert "Expected cost" in [cell.value for cell in resources[1]]

        deployments = load_workbook(
            io.BytesIO((await client.get(f"{API}/exports/deployments.xlsx")).content)
        ).active
        headers = [cell.value for cell in deployments[1]]
        assert "Cost rate" in headers
        assert "Bill rate" not in headers

    async def test_the_billing_export_labels_projected_rows(self, as_role):
        client, _ = await as_role(Role.SALES)
        sheet = load_workbook(
            io.BytesIO((await client.get(f"{API}/exports/billing.xlsx")).content)
        ).active

        # A projected row in a spreadsheet must not read as earned revenue.
        assert "Status" in [cell.value for cell in sheet[1]]

    async def test_exporting_is_audited(self, as_role):
        client, _ = await as_role(Role.SALES)
        await client.get(f"{API}/exports/customers.xlsx")

        admin, _ = await as_role(Role.ADMIN)
        entries = (await admin.get(f"{API}/audit?action=EXPORT_GENERATED")).json()
        assert entries["items"]


# ------------------------------------------------------------ notifications


class TestSlaMilestones:
    @pytest.mark.parametrize(
        ("hours", "expected"),
        [(72, None), (48, 48), (30, 48), (24, 24), (10, 24), (8, 8), (3, 8), (2, 2), (0.5, 2)],
    )
    def test_the_tightest_milestone_reached_is_the_one_that_fires(self, hours, expected):
        assert sla_milestone(hours) == expected

    def test_a_passed_deadline_produces_no_milestone(self):
        assert sla_milestone(-1) is None

    def test_a_missed_sweep_still_fires_the_milestone_it_passed(self):
        # An hourly sweep that skips a run must not lose the 24h alert.
        assert sla_milestone(23) == 24
        assert SLA_MILESTONES == [48, 24, 8, 2]


@pytest.fixture
async def with_alerts(as_role):
    """A requirement whose SLA is inside the window, plus an overdue action."""
    sales, sales_user = await as_role(Role.SALES)

    deadline = datetime.now(UTC) + timedelta(hours=20)
    requirement = (
        await sales.post(
            f"{API}/requirements",
            json={
                "title": f"Sweep Requirement {uuid.uuid4().hex[:6]}",
                "role": "Developer",
                "positions": 1,
                "priority_source": "P5_VENDOR_MSP_VMS",
                "response_deadline_at": deadline.isoformat(),
            },
        )
    ).json()

    opportunity = (
        await sales.post(f"{API}/opportunities", json={"requirement_id": requirement["id"]})
    ).json()
    await sales.patch(
        f"{API}/opportunities/{opportunity['id']}",
        json={
            "next_action": "Call procurement",
            "next_action_due_at": (datetime.now(UTC) - timedelta(days=2)).isoformat(),
        },
    )
    return sales, sales_user, requirement


class TestNotificationSweeps:
    async def test_the_sla_sweep_raises_for_an_approaching_deadline(self, with_alerts, as_role):
        _, _, requirement = with_alerts

        admin, _ = await as_role(Role.ADMIN)
        result = (await admin.post(f"{API}/notifications/sweep")).json()
        assert result["submission_sla"]["raised"] >= 1

        sales, _ = await as_role(Role.SALES)
        inbox = (await sales.get(f"{API}/notifications?category=SUBMISSION_SLA")).json()
        assert any(item["entity_id"] == requirement["id"] for item in inbox)

    async def test_the_follow_up_sweep_catches_a_slipped_action(self, with_alerts, as_role):
        admin, _ = await as_role(Role.ADMIN)
        result = (await admin.post(f"{API}/notifications/sweep")).json()
        assert result["follow_up_overdue"]["raised"] >= 1

    async def test_a_fact_alerts_once(self, with_alerts, as_role):
        """The rule the whole notification system depends on."""
        admin, _ = await as_role(Role.ADMIN)

        first = (await admin.post(f"{API}/notifications/sweep")).json()
        second = (await admin.post(f"{API}/notifications/sweep")).json()

        assert first["submission_sla"]["raised"] >= 1
        assert second["submission_sla"]["raised"] == 0
        assert second["follow_up_overdue"]["raised"] == 0

    async def test_only_an_administrator_can_force_a_sweep(self, as_role):
        client, _ = await as_role(Role.SALES)
        assert (await client.post(f"{API}/notifications/sweep")).status_code == 403


class TestNotificationInbox:
    async def test_the_inbox_is_scoped_to_the_caller(self, with_alerts, as_role):
        admin, _ = await as_role(Role.ADMIN)
        await admin.post(f"{API}/notifications/sweep")

        # SLA alerts target SALES; Resourcing must not see them.
        hr, _ = await as_role(Role.HR_RESOURCING)
        inbox = (await hr.get(f"{API}/notifications?category=SUBMISSION_SLA")).json()
        assert inbox == []

    async def test_the_unread_count_breaks_down_by_category(self, with_alerts, as_role):
        admin, _ = await as_role(Role.ADMIN)
        await admin.post(f"{API}/notifications/sweep")

        sales, _ = await as_role(Role.SALES)
        count = (await sales.get(f"{API}/notifications/unread-count")).json()

        assert count["total"] >= 1
        assert "SUBMISSION_SLA" in count["by_category"]

    async def test_marking_one_read_reduces_the_count(self, with_alerts, as_role):
        admin, _ = await as_role(Role.ADMIN)
        await admin.post(f"{API}/notifications/sweep")

        sales, _ = await as_role(Role.SALES)
        before = (await sales.get(f"{API}/notifications/unread-count")).json()["total"]
        first = (await sales.get(f"{API}/notifications?unread_only=true")).json()[0]

        marked = (await sales.post(f"{API}/notifications/{first['id']}/read")).json()
        assert marked["is_read"] is True
        assert marked["read_at"] is not None

        after = (await sales.get(f"{API}/notifications/unread-count")).json()["total"]
        assert after == before - 1

    async def test_read_all_clears_the_inbox(self, with_alerts, as_role):
        admin, _ = await as_role(Role.ADMIN)
        await admin.post(f"{API}/notifications/sweep")

        sales, _ = await as_role(Role.SALES)
        await sales.post(f"{API}/notifications/read-all")
        assert (await sales.get(f"{API}/notifications/unread-count")).json()["total"] == 0

    async def test_someone_elses_notification_is_a_404_not_a_403(self, with_alerts, as_role):
        """Revealing that another person's alert exists is itself a small leak."""
        admin, _ = await as_role(Role.ADMIN)
        await admin.post(f"{API}/notifications/sweep")

        sales, _ = await as_role(Role.SALES)
        mine = (await sales.get(f"{API}/notifications?category=SUBMISSION_SLA")).json()[0]

        hr, _ = await as_role(Role.HR_RESOURCING)
        assert (await hr.post(f"{API}/notifications/{mine['id']}/read")).status_code == 404

    async def test_every_role_has_an_inbox(self, as_role):
        for role in (Role.ADMIN, Role.MANAGEMENT, Role.SALES, Role.HR_RESOURCING):
            client, _ = await as_role(role)
            assert (await client.get(f"{API}/notifications")).status_code == 200
