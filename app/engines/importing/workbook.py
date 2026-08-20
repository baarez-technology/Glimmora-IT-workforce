"""Reading and writing workbooks.

Deliberately tolerant on the way in and strict on the way out: users send
whatever Excel gave them, and what we hand back has to be unambiguous.

CSV is accepted alongside XLSX because a surprising amount of real staffing data
arrives as a CSV export from somebody else's system.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.engines.importing.schema import Column, EntitySchema, normalise_header

MAX_ROWS = 5000


class WorkbookError(ValueError):
    """The file could not be read at all."""


@dataclass(slots=True)
class ParsedSheet:
    #: Header key -> column index, after normalisation.
    headers: dict[str, int]
    #: (spreadsheet row number, {header key: raw value})
    rows: list[tuple[int, dict[str, Any]]]
    unknown_headers: list[str]
    missing_required: list[str]


def _cells_from_xlsx(payload: bytes) -> list[list[Any]]:
    try:
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises a wide variety of types
        raise WorkbookError("That file could not be opened as a spreadsheet.") from exc

    sheet = workbook.active
    if sheet is None:
        raise WorkbookError("The workbook has no sheets.")
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def _cells_from_csv(payload: bytes) -> list[list[Any]]:
    try:
        # BOM-tolerant: Excel's CSV export writes one, and it otherwise ends up
        # glued to the first header.
        text = payload.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = payload.decode("latin-1")
    return [list(row) for row in csv.reader(io.StringIO(text))]


def parse(payload: bytes, *, filename: str, schema: EntitySchema) -> ParsedSheet:
    """Read a file into normalised rows, without validating any values."""
    lowered = filename.lower()
    if lowered.endswith(".csv"):
        cells = _cells_from_csv(payload)
    elif lowered.endswith((".xlsx", ".xlsm")):
        cells = _cells_from_xlsx(payload)
    else:
        raise WorkbookError("Upload an .xlsx or .csv file.")

    cells = [row for row in cells if any(value not in (None, "") for value in row)]
    if not cells:
        raise WorkbookError("The file is empty.")

    header_row = cells[0]
    headers: dict[str, int] = {}
    unknown: list[str] = []
    known = {column.key for column in schema.columns}

    for index, value in enumerate(header_row):
        key = normalise_header(value)
        if not key:
            continue
        if key in known:
            # First occurrence wins; a duplicated column is reported, not merged.
            headers.setdefault(key, index)
        else:
            unknown.append(str(value).strip())

    missing = [
        column.label for column in schema.columns if column.required and column.key not in headers
    ]

    body = cells[1 : MAX_ROWS + 1]
    rows: list[tuple[int, dict[str, Any]]] = []
    for offset, row in enumerate(body, start=1):
        record = {key: (row[index] if index < len(row) else None) for key, index in headers.items()}
        # Row 1 is the header, so the first data row is row 2 in Excel's own
        # numbering — which is what a user will be looking at.
        rows.append((offset + 1, record))

    return ParsedSheet(
        headers=headers, rows=rows, unknown_headers=unknown, missing_required=missing
    )


# ------------------------------------------------------------------ writing

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ERROR_FILL = PatternFill("solid", fgColor="FEE2E2")
_WARNING_FILL = PatternFill("solid", fgColor="FEF3C7")


def _cell_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime | date):
        return value.isoformat()
    if isinstance(value, list | tuple):
        return "; ".join(str(item) for item in value)
    if isinstance(value, dict):
        return str(value)
    return value


def _autosize(sheet: Any, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(max(width + 2, 10), 60)


def write_sheet(
    *, title: str, headers: list[str], rows: list[list[Any]], freeze: bool = True
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31] or "Sheet1"

    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    widths = [len(str(header)) for header in headers]
    for row in rows:
        values = [_cell_value(value) for value in row]
        sheet.append(values)
        for index, value in enumerate(values):
            widths[index] = max(widths[index], len(str(value or "")))

    _autosize(sheet, widths)
    if freeze:
        sheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def write_error_report(*, schema: EntitySchema, failed: list[dict[str, Any]]) -> bytes:
    """The failed rows, annotated, in a file the user can fix and re-upload.

    Keeping the original columns means the corrected file goes straight back
    through the importer — a plain error list would make them rebuild it.
    """
    columns: list[Column] = schema.columns
    headers = ["Row", "Problem", *[column.label for column in columns]]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Errors"
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    widths = [len(header) for header in headers]
    for entry in failed:
        raw = entry.get("raw") or {}
        problems = "; ".join(entry.get("errors") or entry.get("warnings") or [])
        values = [
            entry.get("row_number"),
            problems,
            *[_cell_value(raw.get(column.key)) for column in columns],
        ]
        sheet.append(values)

        fill = _ERROR_FILL if entry.get("errors") else _WARNING_FILL
        for cell in sheet[sheet.max_row]:
            cell.fill = fill

        for index, value in enumerate(values):
            widths[index] = max(widths[index], len(str(value or "")))

    _autosize(sheet, widths)
    sheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def write_template(schema: EntitySchema) -> bytes:
    """A blank workbook with the right headers and one hint row."""
    headers = [f"{column.label}*" if column.required else column.label for column in schema.columns]
    hints = [column.hint or ("required" if column.required else "") for column in schema.columns]
    return write_sheet(title=schema.entity.value, headers=headers, rows=[hints])


__all__ = [
    "MAX_ROWS",
    "ParsedSheet",
    "WorkbookError",
    "parse",
    "write_error_report",
    "write_sheet",
    "write_template",
]
